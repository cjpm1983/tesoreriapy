import os
from _ast import arg

from django.shortcuts import render, redirect, resolve_url, get_object_or_404, get_list_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import admin

from core.settings import BASE_DIR
from .models import Iglesias, Obreros, Aportesobreros, Aportesiglesias
from django.views import generic
from django.db.models import Sum
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import numbers, NumberFormatDescriptor

# Create your views here.

@login_required(login_url="/admin/login/")
def estadistica(request):
    ###*** Agregar el contexto deladmin es fundamental para que la vista
    # se integre correctamente co lospermisos y lasvariables deladmin
    context = admin.site.each_context(request)
    if 'tipo' and 'item' in request.POST:

        tipo = request.POST['tipo']
        item = request.POST['item']
        if tipo == 'obrero':
            obrero = get_object_or_404(Obreros,id = item)
            aportes = get_list_or_404(Aportesobreros.objects.order_by('anio_id'), obrero_id=item)
            context.__setitem__('obrero', obrero)
            context.__setitem__('aportes', aportes)
            meses = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']
            total = 0
            for mes in meses:
                total += Aportesobreros.objects.filter(obrero_id=item).aggregate(Sum(mes))[mes+'__sum']
            context.__setitem__('total',"$ %.2f" % total)

            return render(request,'tesoreria/obreroDetail.html',context)

        elif tipo == 'iglesia':
            iglesia = get_object_or_404(Iglesias, id=item)
            aportes = get_list_or_404(Aportesiglesias.objects.order_by('anio_id'), iglesia_id=item)
            context.__setitem__('iglesia', iglesia)
            context.__setitem__('aportes', aportes)
            meses = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']
            total = 0
            for mes in meses:
                total += Aportesiglesias.objects.filter(iglesia_id=item).aggregate(Sum(mes))[mes + '__sum']
            context.__setitem__('total', "$ %.2f" % total)
            return render(request, 'tesoreria/iglesiaDetail.html', context)


    elif 'tipo' in request.POST:
        context.__setitem__('tipo', request.POST['tipo'])
        t = request.POST['tipo']
        if t == 'obrero':
            obreros = Obreros.objects.all()
            context.__setitem__('items', obreros)
        elif t == 'iglesia':
            iglesias = Iglesias.objects.all()
            context.__setitem__('items', iglesias)

    return render(request, 'tesoreria/estadistica.html',context)

# class ObreroView(generic.ListView):
#     template_name = 'encuestas/index.html'
#     context_object_name = 'Questions'
#
#     def get_queryset(self):
#         """Return the last five published questions."""
#         return Question.objects.order_by('-pub_date')[:5]

#@login_required(login_url="/admin/login/")
class DetailView(generic.DetailView):
    model = Obreros
    template_name = 'tesoreria/obreroDetail.html'


@login_required(login_url="/admin/login/")
def exportar(request):
    #def exportar(request, tipo, id):
    id = request.POST['id']
    tipo= request.POST['tipo']
    ente = None
    entename = ''
    if tipo == 'obrero':
        ente = get_object_or_404(Obreros, id=id)
        aportes = get_list_or_404(Aportesobreros.objects.order_by('anio_id'), obrero_id=id)
        entename = ente.nombre
    elif tipo == 'iglesia':
        ente = get_object_or_404(Iglesias, id=id)
        aportes = get_list_or_404(Aportesiglesias.objects.order_by('anio_id'), iglesia_id=id)
        entename = ente.iglesia
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml',)
    response['Content-Disposition'] = 'attachment; filename={date}-obreros.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'))

    #renombrar la hoja
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, 'tesoreria/static/tesoreria/plantilla1.xlsx'))
    ws = wb.active
    ws.title = entename

    ws.cell(2,2,'Nombre: %s' % entename)

    if tipo == 'obrero':
        ws.cell(2, 5, 'Folio %s' % str(ente.folio))
        #ws.cell(3, 1, 'Iglesia %s' %str(ente.iglesia_id)) if ente.iglesia_id!=None else "sin definir"

    elif tipo == 'iglesia':
        ws.cell(2, 5, 'Presbiterio: %s' % str(ente.presbiterio))
        ws.cell(2, 8, 'Provincia: %s' % str(ente.provincia))


    #Llenar los mesesa partir de la fila 5
    r=5
    for aporte in aportes:
        ws.cell(r, 1, str(aporte.anio_id))
        ws.cell(r, 2, '=SUM(C%d:N%d)' % (r,r))
        ws.cell(r, 3, float(aporte.ene))
        ws.cell(r, 4, float(aporte.feb))
        ws.cell(r, 5, float(aporte.mar))
        ws.cell(r, 6, float(aporte.abr))
        ws.cell(r, 7, float(aporte.may))
        ws.cell(r, 8, float(aporte.jun))
        ws.cell(r, 9, float(aporte.jul))
        ws.cell(r, 10, float(aporte.ago))
        ws.cell(r, 11, float(aporte.sep))
        ws.cell(r, 12, float(aporte.oct))
        ws.cell(r, 13, float(aporte.nov))
        ws.cell(r, 14, float(aporte.dic))
        r+=1

    #dar formato moneda
    for row in ws.iter_rows(min_row=5,max_row=ws.max_row):
        for cell in row:
            cell.number_format = '[$$-409]#,##0.00;-[$$-409]#,##0.00'


    wb.save(response)
    return response